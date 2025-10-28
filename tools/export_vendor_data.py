#!/usr/bin/env python3
"""Utility to normalize vendor sheets from an XLSX workbook.

The script reads every sheet that resembles a vendor ledger (columns: COD,
Cliente/Cliente2, Localidad plus per-product columns) and generates two
artifacts per sheet:

* ``<seller>_<version>.json`` – Structured JSON payload grouping all clients
  and their purchased products.
* ``<seller>_<version>.csv`` – Flattened CSV (one row per purchased product).

Optionally, the results can be uploaded to an S3 bucket so that they are
available for the ten devices that consume the information. A manifest file is
also generated so clients can detect new weekly versions without inspecting each
individual export.
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Sequence

from openpyxl import load_workbook


@dataclass
class ProductColumn:
    index: int
    code: str


@dataclass
class ClientRow:
    code: str
    name: str
    locality: str
    products: List[str]


@dataclass
class SellerExport:
    seller: str
    json_path: Path
    csv_path: Path
    client_count: int
    product_count: int


def clean(cell_value: object) -> str:
    if cell_value is None:
        return ""
    text = str(cell_value).strip()
    return text


def looks_like_vendor_sheet(headers: Sequence[str]) -> bool:
    has_client = any(h.lower() in {"cliente", "cliente2"} for h in headers)
    has_locality = any(h.lower() == "localidad" for h in headers)
    has_cod = any(h.lower() == "cod" for h in headers)
    return has_client and has_locality and has_cod


def extract_product_columns(headers: Sequence[str]) -> List[ProductColumn]:
    product_columns: List[ProductColumn] = []
    skip = {"cod", "cliente", "cliente2", "localidad", "zona"}
    for idx, header in enumerate(headers):
        header_clean = header.strip()
        if not header_clean:
            continue
        if header_clean.lower() in skip:
            continue
        parts = [part.strip() for part in header_clean.split("-", maxsplit=1)]
        code = parts[0]
        product_columns.append(ProductColumn(index=idx, code=code))
    return product_columns


def parse_vendor_sheet(sheet) -> Iterable[ClientRow]:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [clean(value) for value in header_row]
    if not looks_like_vendor_sheet(headers):
        return []

    idx_cod = headers.index(next(h for h in headers if h.lower() == "cod"))
    idx_client = next(
        i for i, h in enumerate(headers) if h.lower() in {"cliente", "cliente2"}
    )
    idx_localidad = headers.index(next(h for h in headers if h.lower() == "localidad"))

    products = extract_product_columns(headers)

    rows: List[ClientRow] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        code = clean(row[idx_cod]) if idx_cod < len(row) else ""
        if not code:
            continue
        client_name = clean(row[idx_client]) if idx_client < len(row) else ""
        locality = clean(row[idx_localidad]) if idx_localidad < len(row) else ""
        purchased: List[str] = []
        for product in products:
            value = row[product.index] if product.index < len(row) else None
            if clean(value) == "1":
                purchased.append(product.code)
        rows.append(ClientRow(code=code, name=client_name, locality=locality, products=purchased))
    return rows


def compute_week_version(reference: dt.date | None = None) -> str:
    reference = reference or dt.date.today()
    iso_year, iso_week, _ = reference.isocalendar()
    return f"{iso_year}-W{iso_week:02d}"


def export_json(path: Path, seller: str, version: str, rows: Sequence[ClientRow]) -> Path:
    payload = {
        "seller": seller,
        "version": version,
        "generated_at": dt.datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "clients": [
            {
                "code": row.code,
                "name": row.name or row.code,
                "locality": row.locality,
                "products": row.products,
            }
            for row in rows
        ],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def export_csv(path: Path, seller: str, version: str, rows: Sequence[ClientRow]) -> Path:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["version", "seller", "client_code", "client_name", "locality", "product_code"])
        for row in rows:
            if row.products:
                for product_code in row.products:
                    writer.writerow([version, seller, row.code, row.name or row.code, row.locality, product_code])
            else:
                writer.writerow([version, seller, row.code, row.name or row.code, row.locality, ""])
    return path


def upload_to_s3(directory: Path, bucket: str, prefix: str | None = None) -> None:
    import boto3

    s3 = boto3.client("s3")
    for file_path in directory.glob("**/*"):
        if not file_path.is_file():
            continue
        relative_key = file_path.relative_to(directory).as_posix()
        key = f"{prefix.rstrip('/')}/{relative_key}" if prefix else relative_key
        s3.upload_file(str(file_path), bucket, key)


def process_workbook(workbook_path: Path, output_dir: Path, version: str) -> List[SellerExport]:
    wb = load_workbook(workbook_path, data_only=True)
    exports: List[SellerExport] = []
    for sheet in wb.worksheets:
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [clean(value) for value in header_row]
        if not looks_like_vendor_sheet(headers):
            continue
        seller = sheet.title.strip().replace(" ", "_")
        rows = list(parse_vendor_sheet(sheet))
        if not rows:
            continue
        seller_dir = output_dir / seller
        seller_dir.mkdir(parents=True, exist_ok=True)
        json_path = seller_dir / f"{seller}_{version}.json"
        csv_path = seller_dir / f"{seller}_{version}.csv"
        export_json(json_path, seller, version, rows)
        export_csv(csv_path, seller, version, rows)
        exports.append(
            SellerExport(
                seller=seller,
                json_path=json_path,
                csv_path=csv_path,
                client_count=len(rows),
                product_count=sum(len(row.products) for row in rows),
            )
        )
    return exports


def write_manifest(output_dir: Path, version: str, exports: Sequence[SellerExport]) -> Path:
    manifest = {
        "version": version,
        "generated_at": dt.datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "sellers": [
            {
                "seller": export.seller,
                "json": export.json_path.relative_to(output_dir).as_posix(),
                "csv": export.csv_path.relative_to(output_dir).as_posix(),
                "client_count": export.client_count,
                "product_count": export.product_count,
            }
            for export in exports
        ],
    }
    manifest_path = output_dir / f"manifest_{version}.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Normalize vendor sheets from an XLSX workbook.")
    parser.add_argument("workbook", type=Path, help="Path to the vendor XLSX workbook")
    parser.add_argument("--output", type=Path, default=Path("exports"), help="Directory to store generated files")
    parser.add_argument(
        "--version",
        type=str,
        default=compute_week_version(),
        help="Explicit version tag (defaults to current ISO week, e.g. 2024-W09)",
    )
    parser.add_argument("--s3-bucket", type=str, help="Upload results to the given S3 bucket")
    parser.add_argument("--s3-prefix", type=str, help="Optional key prefix when uploading to S3")

    args = parser.parse_args()
    output_dir: Path = args.output
    output_dir.mkdir(parents=True, exist_ok=True)

    exports = process_workbook(args.workbook, output_dir, args.version)
    if not exports:
        print("No vendor sheets were detected in the workbook.")
        return

    manifest_path = write_manifest(output_dir, args.version, exports)
    print(f"Generated exports for {len(exports)} seller(s) under {output_dir}")
    for export in exports:
        print(
            f"  - {export.seller}: {export.client_count} clientes, {export.product_count} productos registrados"
        )
    print(f"Manifest saved to {manifest_path}")

    if args.s3_bucket:
        upload_to_s3(output_dir, args.s3_bucket, args.s3_prefix)
        destination = args.s3_prefix.rstrip("/") if args.s3_prefix else ""
        suffix = f"/{destination}" if destination else ""
        print(f"Uploaded artifacts to s3://{args.s3_bucket}{suffix}")


if __name__ == "__main__":
    main()
